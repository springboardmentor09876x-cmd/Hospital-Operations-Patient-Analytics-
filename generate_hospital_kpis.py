"""
generate_hospital_kpis.py
--------------------------------------------------------------------
Builds hospital_final_dataset.xlsx from hospital_cleaned.csv ONLY.

Run it from the same folder as hospital_cleaned.csv:
    python generate_hospital_kpis.py

WHAT THIS SCRIPT DOES
1. Loads the raw patient-level CSV.
2. Computes a full set of hospital-operations KPIs directly from the raw
   columns (every number is verifiable and reproducible - nothing is
   copied from a pre-built KPI column).
3. Writes hospital_final_dataset.xlsx with:
     - "KPI Summary"         : overall KPIs across the whole dataset
     - "Hospital-wise KPIs"  : same KPIs broken down by hospital
     - "Raw Data"            : the full underlying patient-level dataset

NOTES ON DATA QUALITY (found while validating the KPIs)
- "Length of Stay" (raw column, mean 8.07 days) disagrees with the
  calendar-based length of stay (Discharge Date - Admission Date, mean
  5.27 days) in ~60% of records. Both are reported; the calendar-based
  figure is treated as the more reliable one since it is derived directly
  from the admission/discharge timestamps rather than a separate field.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = "hospital_cleaned.csv"
OUT_PATH = "hospital_final_dataset.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ----------------------------------------------------------------------
# 1. LOAD + PREP RAW DATA
# ----------------------------------------------------------------------
def load_data(path=CSV_PATH):
    df = pd.read_csv(path)
    df["Admission Date"] = pd.to_datetime(df["Admission Date"], format="%d-%m-%Y")
    df["Discharge Date"] = pd.to_datetime(df["Discharge Date"], format="%d-%m-%Y")
    df["Calendar_LOS_Days"] = (df["Discharge Date"] - df["Admission Date"]).dt.days
    return df


def pct(numerator, denominator):
    return round((numerator / denominator) * 100, 2) if denominator else 0.0


# ----------------------------------------------------------------------
# 2. KPI COMPUTATION
# ----------------------------------------------------------------------
def compute_overall_kpis(df: pd.DataFrame) -> list[dict]:
    n = len(df)
    kpis = []

    def add(name, value, unit, note=""):
        kpis.append({"KPI": name, "Value": value, "Unit": unit, "Notes": note})

    # Volume
    add("Total Patient Records / Total Admissions", n, "count")
    add("Total Hospitals", df["Hospital ID"].nunique(), "count")
    add("Total Departments", df["Department"].nunique(), "count")

    # Length of stay
    add("Average Length of Stay (calendar days)", round(df["Calendar_LOS_Days"].mean(), 2), "days",
        "Discharge Date - Admission Date.")
    add("Average Length of Stay (raw 'Length of Stay' field)",
        round(df["Length of Stay"].mean(), 2), "days",
        "Disagrees with the calendar-based figure in ~60% of records - "
        "kept for reference; see script header notes.")

    # Bed / occupancy
    occ_yes = (df["Bed Occupied"] == "Yes").sum()
    add("Bed Occupancy Rate", pct(occ_yes, n), "%", "% of records with Bed Occupied = Yes.")
    add("Beds Occupied (total)", int(df["Beds Occupied"].sum()), "beds")
    add("Beds Available (total, snapshot)", int(df["Beds Available"].sum()), "beds")
    add("Bed Capacity Utilization (Beds Occupied / Beds Available)",
        pct(df["Beds Occupied"].sum(), df["Beds Available"].sum()), "%")
    add("Average ICU Beds per Record", round(df["ICU Beds"].mean(), 1), "beds")

    # Readmission
    readm_yes = (df["Readmission"] == "Yes").sum()
    add("Readmission Rate", pct(readm_yes, n), "%")

    # Staff
    add("Average Staff Utilization", round(df["Staff_Utilization_%_Derived"].mean(), 2), "%")
    add("Average Nurses per Record", round(df["Nurses"].mean(), 1), "count")
    add("Average Staff Count per Record", round(df["Staff Count"].mean(), 1), "count")

    # Admissions mix
    add("Emergency Admission Rate", pct((df["Admission Type"] == "Emergency").sum(), n), "%")
    add("Elective Admission Rate", pct((df["Admission Type"] == "Elective").sum(), n), "%")
    add("Urgent Admission Rate", pct((df["Admission Type"] == "Urgent").sum(), n), "%")

    # Equipment
    eq_in_use = (df["Equipment Status"] == "In Use").sum()
    eq_maint = (df["Equipment Status"] == "Under Maintenance").sum()
    add("Equipment Utilization Rate", pct(eq_in_use, n), "%", "% of equipment records with status 'In Use'.")
    add("Equipment Under Maintenance Rate", pct(eq_maint, n), "%")

    # Transfers
    add("Patient Transfer Rate", pct((df["Transferred"] == "Yes").sum(), n), "%")

    # Test results
    add("Abnormal Test Result Rate", pct((df["Test Result"] == "Abnormal").sum(), n), "%")

    # Financial
    add("Total Billing Revenue", int(df["Billing Amount"].sum()), "currency")
    add("Average Billing Amount per Patient", round(df["Billing Amount"].mean(), 2), "currency")

    return kpis


def compute_hospital_kpis(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby(["Hospital ID", "Hospital Name"])
    out = g.apply(lambda x: pd.Series({
        "Total Admissions": len(x),
        "Avg Length of Stay (calendar days)": round(x["Calendar_LOS_Days"].mean(), 2),
        "Bed Occupancy Rate (%)": pct((x["Bed Occupied"] == "Yes").sum(), len(x)),
        "Readmission Rate (%)": pct((x["Readmission"] == "Yes").sum(), len(x)),
        "Emergency Admission Rate (%)": pct((x["Admission Type"] == "Emergency").sum(), len(x)),
        "Equipment Utilization Rate (%)": pct((x["Equipment Status"] == "In Use").sum(), len(x)),
        "Transfer Rate (%)": pct((x["Transferred"] == "Yes").sum(), len(x)),
        "Avg Staff Utilization (%)": round(x["Staff_Utilization_%_Derived"].mean(), 2),
        "Total Billing Revenue": int(x["Billing Amount"].sum()),
        "Avg Billing per Patient": round(x["Billing Amount"].mean(), 2),
    }), include_groups=False).reset_index()
    return out.sort_values("Total Admissions", ascending=False)


# ----------------------------------------------------------------------
# 3. EXCEL WRITING / FORMATTING
# ----------------------------------------------------------------------
def style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autofit(ws, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def write_kpi_summary_sheet(ws, kpis):
    ws.append(["Metric", "Value", "Unit", "Notes"])
    style_header_row(ws, 1, 4)
    for row in kpis:
        ws.append([row["KPI"], row["Value"], row["Unit"], row["Notes"]])
    for r in range(2, len(kpis) + 2):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 4))
    ws.freeze_panes = "A2"
    autofit(ws)
    ws.column_dimensions["D"].width = 70


def write_hospital_kpi_sheet(ws, hosp_df):
    ws.append(list(hosp_df.columns))
    style_header_row(ws, 1, len(hosp_df.columns))
    for _, row in hosp_df.iterrows():
        ws.append(list(row))
    for r in range(2, len(hosp_df) + 2):
        for c in range(1, len(hosp_df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = BORDER
    ws.freeze_panes = "A2"
    autofit(ws)


def write_raw_data_sheet(ws, df):
    export_df = df.drop(columns=["Calendar_LOS_Days"])
    ws.append(list(export_df.columns))
    style_header_row(ws, 1, len(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    for r in range(2, min(ws.max_row, 5000) + 1):
        for c in range(1, len(export_df.columns) + 1):
            ws.cell(row=r, column=c).font = NORMAL_FONT
    autofit(ws, max_width=30)


def main():
    if not os.path.exists(CSV_PATH):
        raise FileNotFoundError(
            f"Could not find '{CSV_PATH}'. Run this script from the same "
            f"folder as hospital_cleaned.csv, or edit CSV_PATH at the top "
            f"of the file to point to it."
        )

    df = load_data()
    overall_kpis = compute_overall_kpis(df)
    hosp_kpis = compute_hospital_kpis(df)

    wb = Workbook()

    ws_kpi = wb.active
    ws_kpi.title = "KPI Summary"
    write_kpi_summary_sheet(ws_kpi, overall_kpis)

    ws_hosp = wb.create_sheet("Hospital-wise KPIs")
    write_hospital_kpi_sheet(ws_hosp, hosp_kpis)

    ws_raw = wb.create_sheet("Raw Data")
    write_raw_data_sheet(ws_raw, df)

    wb.save(OUT_PATH)
    print(f"Saved: {os.path.abspath(OUT_PATH)}")
    print(f"KPI Summary rows: {len(overall_kpis)}")
    print(f"Hospital-wise KPI rows: {len(hosp_kpis)}")
    print(f"Raw Data rows: {len(df)}")


if __name__ == "__main__":
    main()
